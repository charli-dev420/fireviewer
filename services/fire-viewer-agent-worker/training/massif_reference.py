from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

EXPECTED_SOURCE_BYTES = 5_571_072
EXPECTED_SOURCE_SHA1 = "7f31890d29a03ab2c8e65841296b8d5fc23fe007"
SOURCE_PAGE = "https://www.data.gouv.fr/datasets/les-perimetres-de-massifs"
RESOURCE_URL = (
    "https://static.data.gouv.fr/resources/les-perimetres-de-massifs/"
    "20210907-123330/diffusion-zonages-massifs-cog2021.xls"
)
SHEET_NAME = "Communes de massif (COG 2021)"
CODE_PATTERN = re.compile(r"(?:\d{5}|2[AB]\d{3})")
PARTIAL_SUFFIX = " (partiellement)"
MASSIF_IDS = {
    "Alpes": "alpes",
    "Corse": "corse",
    "Guadeloupe": "guadeloupe",
    "Hauts de la Réunion": "hauts-de-la-reunion",
    "Jura": "jura",
    "Martinique": "martinique",
    "Massif Central": "massif-central",
    "Pyrénées": "pyrenees",
    "Vosges": "vosges",
}


@dataclass(frozen=True)
class MassifCommune:
    code_insee: str
    commune_name: str
    massif_id: str
    massif_name: str
    membership: str


def sha1_file(path: Path) -> str:
    digest = hashlib.sha1(usedforsecurity=False)
    with path.open("rb") as handle:
        while chunk := handle.read(8 * 1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_rows(rows: Iterable[Sequence[object]]) -> list[MassifCommune]:
    records: list[MassifCommune] = []
    seen_codes: set[str] = set()
    for row_number, row in enumerate(rows, start=1):
        if len(row) < 3:
            raise ValueError(f"Massif row {row_number} has fewer than three columns")
        if all(not str(value or "").strip() for value in row[:3]):
            continue
        code = str(row[0] or "").strip()
        if code.isdigit():
            code = code.zfill(5)
        commune_name = str(row[1] or "").strip()
        raw_massif = str(row[2] or "").strip()
        if raw_massif == "Hors massif":
            continue
        membership = "partial" if raw_massif.endswith(PARTIAL_SUFFIX) else "full"
        massif_name = raw_massif.removesuffix(PARTIAL_SUFFIX)
        if not CODE_PATTERN.fullmatch(code):
            raise ValueError(f"Invalid COG 2021 code at row {row_number}: {code!r}")
        if not commune_name:
            raise ValueError(f"Missing commune name at row {row_number}")
        if massif_name not in MASSIF_IDS:
            raise ValueError(f"Unknown massif at row {row_number}: {raw_massif!r}")
        if code in seen_codes:
            raise ValueError(f"Duplicate COG 2021 code: {code}")
        seen_codes.add(code)
        records.append(
            MassifCommune(
                code_insee=code,
                commune_name=commune_name,
                massif_id=MASSIF_IDS[massif_name],
                massif_name=massif_name,
                membership=membership,
            )
        )
    return records


def _xlsx_rows(path: Path) -> Iterable[Sequence[object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[SHEET_NAME]
        iterator = sheet.iter_rows(values_only=True)
        for row in iterator:
            normalized = tuple(str(value or "").strip() for value in row[:3])
            if normalized == ("CODGEO", "LIBGEO", "MASSIF"):
                break
        else:
            raise ValueError(f"Header CODGEO/LIBGEO/MASSIF not found in {SHEET_NAME!r}")
        yield from iterator
    finally:
        workbook.close()


def _write_csv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, Any]]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(path)


def build_reference(input_xlsx: Path, source_xls: Path, output_dir: Path) -> dict[str, Any]:
    if source_xls.stat().st_size != EXPECTED_SOURCE_BYTES:
        raise ValueError("Official massif XLS size does not match the pinned resource")
    source_sha1 = sha1_file(source_xls)
    if source_sha1 != EXPECTED_SOURCE_SHA1:
        raise ValueError("Official massif XLS SHA-1 does not match the pinned resource")

    records = normalize_rows(_xlsx_rows(input_xlsx))
    if len(records) != 8_617:
        raise ValueError(f"Expected 8,617 in-massif communes, found {len(records)}")
    output_dir.mkdir(parents=True, exist_ok=True)
    _write_csv(
        output_dir / "communes_massifs_cog2021.csv",
        ["code_insee", "commune_name", "massif_id", "massif_name", "membership"],
        (asdict(record) for record in records),
    )

    full_counts = Counter(record.massif_name for record in records if record.membership == "full")
    partial_counts = Counter(
        record.massif_name for record in records if record.membership == "partial"
    )
    summaries = [
        {
            "massif_id": MASSIF_IDS[name],
            "massif_name": name,
            "full_communes": full_counts[name],
            "partial_communes": partial_counts[name],
            "total_communes": full_counts[name] + partial_counts[name],
        }
        for name in sorted(MASSIF_IDS)
    ]
    _write_csv(
        output_dir / "massifs_summary_cog2021.csv",
        [
            "massif_id",
            "massif_name",
            "full_communes",
            "partial_communes",
            "total_communes",
        ],
        summaries,
    )

    report = {
        "source_id": "anct_massifs_cog2021",
        "source_page": SOURCE_PAGE,
        "resource_url": RESOURCE_URL,
        "resource_bytes": EXPECTED_SOURCE_BYTES,
        "resource_sha1": source_sha1,
        "license": "Licence-Ouverte-2.0",
        "reference_version": "COG-2021",
        "raw_commune_rows": 34_965,
        "kept_in_massif_rows": len(records),
        "excluded_hors_massif_rows": 26_348,
        "massif_count": len(summaries),
        "usage_constraint": (
            "Commune membership table only; join to a separately pinned COG 2021 commune "
            "geometry before any spatial point-in-polygon operation."
        ),
    }
    (output_dir / "provenance.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the pinned COG 2021 massif reference")
    parser.add_argument("--input-xlsx", type=Path, required=True)
    parser.add_argument("--source-xls", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    print(
        json.dumps(
            build_reference(args.input_xlsx, args.source_xls, args.output),
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
